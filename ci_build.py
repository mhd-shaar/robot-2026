# -*- coding: utf-8 -*-
"""يُشغَّل داخل GitHub Actions: يسحب البيانات من المحرّر المستضاف ويحدّث followup.html.

لا يلمس صفحات البرنامج ولا الرئيسية — تلك تُولَّد من جهاز الإدارة.
لا يكتب شيئًا إن لم تتغيّر البيانات (لتفادي تاريخ مليء بتعديلات فارغة).

    python ci_build.py     →  يعيد 0 دائمًا؛ يطبع ما إذا تغيّر شيء
"""
import openpyxl, html as _h, os, io, hashlib, urllib.request
from datetime import datetime, timezone, timedelta

EDITOR = "https://shaar.pythonanywhere.com"
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "followup.html")
HASHFILE = os.path.join(HERE, ".data-hash")
MAXCOL = 64

# ─────────────────────────── سحب البيانات ───────────────────────────
req = urllib.request.Request(EDITOR + "/export.xlsx", headers={"User-Agent": "robot-2026-ci"})
data = urllib.request.urlopen(req, timeout=120).read()
# data_only=True مطابقةً للناشر المحلّي: ورقة الكوادر فيها معادلات COUNTA
wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
if not wb.sheetnames:
    raise SystemExit("الملف المسحوب لا يحوي أوراقًا — أُلغيت العملية")

# بصمة على القيم فقط (لا على بايتات الملف، فهي تتغيّر عند كل تصدير)
sig = hashlib.sha256()
for ws in wb.worksheets:
    sig.update(ws.title.encode())
    for row in ws.iter_rows(max_col=min(ws.max_column or 1, MAXCOL)):
        for c in row:
            if c.value is not None:
                sig.update(f"{c.coordinate}={c.value}".encode())
digest = sig.hexdigest()

old = open(HASHFILE, encoding="utf-8").read().strip() if os.path.exists(HASHFILE) else ""
if old == digest:
    print("لا تغيير في البيانات — لا حاجة للتحديث")
    raise SystemExit(0)

# ─────────────────────────── التصيير ───────────────────────────
def rgb(c):
    try:
        v = c.rgb
        if isinstance(v, str) and len(v) in (6, 8):
            return "#" + v[-6:]
    except Exception:
        pass
    return None

def bounds(ws):
    cap = min(ws.max_column or 1, MAXCOL)
    mr = mc = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=cap):
        for c in row:
            if c.value is not None:
                mr = max(mr, c.row); mc = max(mc, c.column)
    for m in ws.merged_cells.ranges:
        if m.max_col <= cap:
            mr = max(mr, m.max_row); mc = max(mc, m.max_col)
    return mr, mc

def table(ws):
    mr, mc = bounds(ws)
    if mr == 0:
        return '<p class="muted">(ورقة فارغة)</p>'
    span, skip = {}, set()
    for m in ws.merged_cells.ranges:
        span[(m.min_row, m.min_col)] = (m.max_row-m.min_row+1, m.max_col-m.min_col+1)
        for r in range(m.min_row, m.max_row+1):
            for c in range(m.min_col, m.max_col+1):
                if (r, c) != (m.min_row, m.min_col):
                    skip.add((r, c))
    cols = ""
    for ci in range(1, mc+1):
        L = openpyxl.utils.get_column_letter(ci)
        w = ws.column_dimensions[L].width if L in ws.column_dimensions else None
        cols += f'<col style="width:{round((w or 12)*8)}px">'
    out = f'<table><colgroup>{cols}</colgroup><tbody>'
    for r in range(1, mr+1):
        out += "<tr>"
        for ci in range(1, mc+1):
            if (r, ci) in skip:
                continue
            cell = ws.cell(r, ci); st = []
            if cell.fill and cell.fill.patternType:
                bg = rgb(cell.fill.fgColor)
                if bg:
                    st.append(f"background:{bg}")
            f = cell.font
            if f:
                fc = rgb(f.color) if f.color else None
                if fc: st.append(f"color:{fc}")
                if f.bold: st.append("font-weight:800")
                if f.size: st.append(f"font-size:{float(f.size)}px")
            if cell.alignment and cell.alignment.horizontal:
                st.append(f"text-align:{cell.alignment.horizontal}")
            rs, cs = span.get((r, ci), (1, 1))
            a = (f' rowspan="{rs}"' if rs > 1 else "") + (f' colspan="{cs}"' if cs > 1 else "")
            v = "" if cell.value is None else _h.escape(str(cell.value))
            out += f'<td{a} style="{";".join(st)}">{v}</td>'
        out += "</tr>"
    return out + "</tbody></table>"

# نفس تنسيقات الموقع (يجب أن تبقى مطابقة لما في ناشر_الموقع.py)
CSS = """
*{margin:0;padding:0;box-sizing:border-box}
:root{--navy:#12315a;--gold:#b8912f;--gold2:#d8b559;--line:#e3dcc7;--paper:#f4f2ea}
body{font-family:"Segoe UI","Tahoma",Arial,sans-serif;background:var(--paper);color:#1c2430}
.nav{background:var(--navy);padding:11px 18px;display:flex;align-items:center;gap:14px;flex-wrap:wrap}
.nav img{height:34px}
.nav a{color:#dfe6ef;text-decoration:none;font-size:13.5px;font-weight:700;padding:7px 13px;border-radius:7px}
.nav a:hover{background:#1d4577;color:#fff}
.nav a.on{background:var(--gold);color:#fff}
.nav .sp{flex:1}
.wrap{max-width:1180px;margin:0 auto;padding:20px 16px 50px}
h1{font-size:22px;color:var(--navy)} h2{font-size:16px;color:var(--navy);margin-bottom:9px}
.muted{font-size:12px;color:#7a7663}
.tabs{display:flex;flex-wrap:wrap;gap:6px;margin:14px 0 11px}
.tab{font:700 13px inherit;padding:8px 15px;border:1px solid #ddd6c2;background:#fff;color:#42505f;
 border-radius:8px 8px 0 0;cursor:pointer}
.tab.active{background:var(--navy);color:#fff;border-color:var(--navy)}
.pane{display:none} .pane.active{display:block}
.scroll{overflow-x:auto}
table{border-collapse:collapse;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.09)}
td{border:1px solid #d9d3c2;padding:5px 9px;font-size:12.5px;vertical-align:middle;
 white-space:pre-wrap;word-break:break-word}
.foot{text-align:center;font-size:11.5px;color:#8a8676;margin-top:28px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:14px;margin-top:20px}
.card{background:#fff;border:1px solid var(--line);border-radius:12px;padding:20px;text-decoration:none;
 color:inherit;display:block;transition:.15s}
.card:hover{transform:translateY(-2px);box-shadow:0 5px 16px rgba(0,0,0,.11);border-color:var(--gold2)}
.card .ic{font-size:28px} .card h3{font-size:16px;color:var(--navy);margin:8px 0 5px}
.card p{font-size:12.5px;color:#6b7480;line-height:1.5}
.hero{text-align:center;padding:26px 0 6px}
.hero img{height:82px} .hero .k{font-size:10.5px;letter-spacing:3px;color:var(--gold);font-weight:700;margin-top:9px}
.hero h1{font-size:27px;margin:4px 0} .hero .d{font-size:13.5px;color:#5a6472;font-weight:600}
.ro{display:inline-block;background:#eef1f6;color:var(--navy);border:1px solid #ccd7e6;
 border-radius:20px;padding:5px 14px;font-size:11.5px;font-weight:700;margin-top:10px}
.split{display:flex;align-items:center;gap:14px;margin:34px 0 4px;color:var(--navy);
 font-weight:800;font-size:15px}
.split::before,.split::after{content:"";flex:1;height:1px;background:var(--line)}
.lead{font-size:13px;color:#5a6472;line-height:1.6;text-align:center;margin-bottom:2px}
.days{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px;margin-top:14px}
.day{background:#fff;border:1px solid var(--line);border-radius:12px;padding:17px 19px;
 border-top:4px solid var(--gold2)}
.day.on{border-top-color:var(--navy)}
.day .dt{display:inline-block;background:#eef1f6;color:var(--navy);border-radius:14px;
 padding:3px 12px;font-size:11.5px;font-weight:800}
.day>b{display:block;font-size:15.5px;color:var(--navy);margin:8px 0 6px}
.day p{font-size:12.5px;color:#5a6472;line-height:1.65}
.day p b{color:var(--navy);font-weight:800}
"""

def nav(active):
    it = [("index.html", "الرئيسية"), ("program-general.html", "البرنامج العام"),
          ("program-detailed.html", "البرنامج التفصيلي"), ("followup.html", "قوائم المتابعة")]
    links = ""
    for h, t in it:
        cls = ' class="on"' if h == active else ""
        links += f'<a href="{h}"{cls}>{t}</a>'
    return f'<div class="nav"><img src="logo.png" alt="">{links}<span class="sp"></span></div>'

stamp = (datetime.now(timezone.utc) + timedelta(hours=3)).strftime("%Y-%m-%d  %H:%M")   # توقيت دمشق
tabs = panes = ""
for i, ws in enumerate(wb.worksheets):
    a = " active" if i == 0 else ""
    tabs += f'<button class="tab{a}" onclick="show({i},this)">{_h.escape(ws.title)}</button>'
    panes += f'<div class="pane{a}"><div class="scroll">{table(ws)}</div></div>'

body = (f'<h1>قوائم المتابعة</h1>'
        f'<div class="muted">عرض فقط — المصدر ملف الإكسل لدى إدارة البطولة.</div>'
        f'<div class="tabs">{tabs}</div>{panes}'
        '<script>function show(i,el){'
        'document.querySelectorAll(".pane").forEach(function(p,j){p.classList.toggle("active",j===i)});'
        'document.querySelectorAll(".tab").forEach(function(t){t.classList.remove("active")});'
        'el.classList.add("active");}</script>')

html = f"""<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>قوائم المتابعة — بطولة الروبوت 2026</title><link rel="icon" href="logo.png"><style>{CSS}</style></head><body>
{nav("followup.html")}
<div class="wrap">{body}
<div class="foot">ثانوية المحدّث الأكبر الشيخ بدر الدين الحسني الشرعية للبنين — بطولة الروبوت 2026
<br>آخر تحديث: {stamp}</div></div></body></html>"""

with open(OUT, "w", encoding="utf-8") as f:
    f.write(html)
with open(HASHFILE, "w", encoding="utf-8") as f:
    f.write(digest)
print(f"تحدّثت القوائم — {len(wb.sheetnames)} أوراق، {sum(ws.max_row for ws in wb.worksheets)} صفًّا")
print("CHANGED=1")
